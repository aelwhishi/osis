##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2018 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from operator import itemgetter

from django.core.management import BaseCommand
from django.db.models import Prefetch
from django.utils import translation
from openpyxl import Workbook

from attribution.models.attribution import Attribution
from base.models.entity_container_year import EntityContainerYear
from base.models.entity_version import get_last_version
from base.models.enums import entity_container_year_link_type, learning_container_year_types
from base.models.learning_unit_year import LearningUnitYear
from base.models.program_manager import ProgramManager


class Command(BaseCommand):
    def handle(self, *args, **options):
        #_extract_program_managers_to_xls()
        _extract_learning_unit_with_score_responsible()


def _extract_program_managers_to_xls():
    qs_list = ProgramManager.objects.filter(
        offer_year__academic_year__year=2018,
    ).order_by(
        'offer_year__entity_management_fac__acronym',
        'offer_year__entity_management__acronym',
        'offer_year__acronym',
        'person__last_name',
        'person__first_name',
    ).select_related(
        'offer_year',
        'person',
        'offer_year__entity_management_fac',
        'offer_year__entity_management',
    )

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["Faculté de l'entité de gestion", 'Entité de gestion', 'Sigle', 'Intitulé', 'Nom', 'Prénom'])
    for program_manager in qs_list:
        row = [
            program_manager.offer_year.entity_management_fac.acronym
            if program_manager.offer_year.entity_management_fac else '',
            program_manager.offer_year.entity_management.acronym
            if program_manager.offer_year.entity_management else '',
            program_manager.offer_year.acronym,
            program_manager.offer_year.title,
            program_manager.person.last_name,
            program_manager.person.first_name
        ]
        worksheet.append(row)

    workbook.save(filename='program_managers_2018.xlsx')


def _extract_learning_unit_with_score_responsible():
    qs_list = LearningUnitYear.objects.filter(
        academic_year__year=2018,
    ).exclude(
        learning_container_year__container_type=learning_container_year_types.EXTERNAL
    ).prefetch_related(
        Prefetch(
            'learning_container_year__entitycontaineryear_set',
            queryset=EntityContainerYear.objects.filter(type=entity_container_year_link_type.REQUIREMENT_ENTITY),
            to_attr='requirement_entity'
        ),
        Prefetch(
            'attribution_set',
            queryset=Attribution.objects.all().select_related('tutor__person'),
            to_attr='attributions'
        )
    )

    with translation.override('fr_BE'):
        xls_dict = {}
        for learning_unit_year in qs_list:
            requirement_entity = None
            fac_requirement_entity = None
            code = learning_unit_year.acronym
            type = learning_unit_year.learning_container_year.get_container_type_display() if \
                learning_unit_year.learning_container_year else "Classe"
            intitule = learning_unit_year.complete_title
            tutors = set()
            score_responsibles = set()
            if learning_unit_year.learning_container_year and learning_unit_year.learning_container_year.requirement_entity:
                entity = learning_unit_year.learning_container_year.requirement_entity[0].entity
                requirement_entity = get_last_version(entity)

                if requirement_entity:
                    fac_requirement_entity = requirement_entity.find_faculty_version(learning_unit_year.academic_year)

            for attribution in learning_unit_year.attributions:
                pattern = "{last_name} {first_name} ({function})"
                tutor_text = pattern.format(
                    last_name=attribution.tutor.person.last_name,
                    first_name=attribution.tutor.person.first_name,
                    function=attribution.get_function_display()
                )
                tutors.add(tutor_text)
                if attribution.score_responsible:
                    score_responsibles.add(tutor_text)

            tutors = sorted(tutors)
            score_responsibles = sorted(score_responsibles)
            xls_dict[code] = (
                fac_requirement_entity.acronym if fac_requirement_entity else '',
                requirement_entity.acronym if requirement_entity else '',
                code,
                type,
                intitule,
                "\n".join(tutors),
                "\n".join(score_responsibles),
            )

    # Resolve classe requirement faculty
    xls_dict_resolved = {code: _resolve_class(tuple, xls_dict) for code, tuple in xls_dict.items()}

    # Sort result
    xls_data = sorted(xls_dict_resolved.values(), key=itemgetter(0, 1, 2))
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["Faculté de l'entité de gestion", 'Entité de gestion', 'Sigle', 'Type', 'Intitulé',
                      'Enseignant(s)', 'Responsable de notes'])
    for idx, row in enumerate(xls_data, 1):
        worksheet.cell('F{}'.format(idx)).style.alignment.wrap_text = True
        worksheet.cell('G{}'.format(idx)).style.alignment.wrap_text = True
        worksheet.append(list(row))
    workbook.save(filename='learning_unit_2018.xlsx')


def _resolve_class(tuple, xls_dict):
    if tuple[3] == 'Classe':
        tuple_parent = xls_dict.get(tuple[2][:-1])
        if tuple_parent:
            return (tuple_parent[0], tuple_parent[1],) + (tuple[2], tuple[3], tuple[4], tuple[5], tuple[6], )
    return tuple
